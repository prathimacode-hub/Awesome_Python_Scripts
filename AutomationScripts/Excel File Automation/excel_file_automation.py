from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

#making the grid and giving the data
data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
#assigning the title
ws.title = "Grades"
#giving proper formatting for columns 
headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

#reading the data for persom
for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['Joe']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"
#assigning the colour and text type 
for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

#saving the excel file in the same folder
wb.save("NewGrades.xlsx")
